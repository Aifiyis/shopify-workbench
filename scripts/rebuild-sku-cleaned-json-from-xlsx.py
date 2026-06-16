import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "storage" / "app" / "private" / "sku-cleaned.xlsx"
JSON_PATH = ROOT / "storage" / "app" / "private" / "sku-cleaned.json"
SHEET_NAME = "原始数据"
HEADERS = ["original_sku", "cleaned_sku", "中文名称", "上品人", "工艺", "处理人"]


def cell_to_string(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {XLSX_PATH}")

    workbook = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Sheet not found: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    header_row = [cell_to_string(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    missing_headers = [header for header in HEADERS if header not in header_row]

    if missing_headers:
        raise RuntimeError(f"Missing required headers: {', '.join(missing_headers)}")

    column_indexes = {header: header_row.index(header) for header in HEADERS}
    rows = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {header: cell_to_string(values[column_indexes[header]] if column_indexes[header] < len(values) else "") for header in HEADERS}

        if row["original_sku"] == "":
            continue

        rows.append(row)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = JSON_PATH.with_name(f"sku-cleaned.before-xlsx-refresh-{timestamp}.json")

    if JSON_PATH.exists():
        shutil.copy2(JSON_PATH, backup_path)

    JSON_PATH.write_text(json.dumps(rows, ensure_ascii=False, indent=4), encoding="utf-8")

    print(json.dumps({
        "source": str(XLSX_PATH),
        "sheet": SHEET_NAME,
        "output": str(JSON_PATH),
        "backup": str(backup_path) if backup_path.exists() else None,
        "rows": len(rows),
        "headers": HEADERS,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
